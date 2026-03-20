"""
01_explore_data.py - Explore and visualize all three battery datasets.
"""
import os, sys
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import scipy.io as sio
import openpyxl
import json

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(WORKSPACE, 'data')
IMG_DIR = os.path.join(WORKSPACE, 'report', 'images')
OUT_DIR = os.path.join(WORKSPACE, 'outputs')
os.makedirs(IMG_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

plt.rcParams.update({'font.size': 11, 'figure.dpi': 150, 'savefig.bbox': 'tight'})

# ============================================================
# 1. CS2_36 Dataset (CALCE)
# ============================================================
print("Loading CS2_36 (CALCE) dataset...")
cs2_dir = os.path.join(DATA_DIR, 'CS2_36')
cs2_files = sorted([f for f in os.listdir(cs2_dir) if f.endswith('.xlsx')])

cs2_channel_data = {}
cs2_stats_data = {}
for f in cs2_files:
    fpath = os.path.join(cs2_dir, f)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if 'Channel' not in sheet_name and 'Statistics' not in sheet_name:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
        data_dict = {}
        for i, h in enumerate(headers):
            vals = []
            for r in rows[1:]:
                if r[i] is not None:
                    try:
                        vals.append(float(r[i]))
                    except (ValueError, TypeError):
                        pass
            data_dict[h] = np.array(vals)
        key = f.replace('.xlsx', '')
        if 'Channel' in sheet_name:
            cs2_channel_data[key] = (headers, data_dict)
        elif 'Statistics' in sheet_name:
            cs2_stats_data[key] = (headers, data_dict)
    wb.close()

print(f"Loaded {len(cs2_channel_data)} channel sheets, {len(cs2_stats_data)} stats sheets")

# Plot CS2_36 discharge curves from channel data
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('CS2_36 (CALCE) - Discharge Voltage Curves at Different Cycles', fontsize=14, fontweight='bold')

for idx, (key, (headers, dd)) in enumerate(cs2_channel_data.items()):
    ax = axes[idx // 2][idx % 2]
    # Extract discharge segments (negative current = discharge)
    time = dd.get('Test_Time(s)', np.array([]))
    voltage = dd.get('Voltage(V)', np.array([]))
    current = dd.get('Current(A)', np.array([]))
    cycle_idx = dd.get('Cycle_Index', np.array([]))

    if len(voltage) > 0 and len(current) > 0:
        # Plot a few discharge cycles
        unique_cycles = np.unique(cycle_idx)
        colors = plt.cm.viridis(np.linspace(0, 1, min(10, len(unique_cycles))))
        plotted = 0
        for ci, c in enumerate(unique_cycles[::max(1, len(unique_cycles)//10)]):
            mask = (cycle_idx == c) & (current < -0.01)
            if np.sum(mask) > 10:
                t_seg = time[mask] - time[mask][0]
                v_seg = voltage[mask]
                ax.plot(t_seg / 3600, v_seg, color=colors[plotted % len(colors)],
                       linewidth=1, label=f'Cycle {int(c)}')
                plotted += 1
                if plotted >= 10:
                    break

    ax.set_xlabel('Time (h)')
    ax.set_ylabel('Voltage (V)')
    ax.set_title(key, fontsize=10)
    ax.legend(fontsize=7, ncol=2)
    ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'cs2_36_overview.png'))
plt.close()
print("Saved cs2_36_overview.png")

# Plot capacity fade for CS2_36
fig, ax = plt.subplots(figsize=(10, 6))
fig.suptitle('CS2_36 - Capacity Fade Across Files', fontsize=14, fontweight='bold')
for key, (headers, dd) in cs2_stats_data.items():
    cycle = dd.get('Cycle_Index', np.array([]))
    discharge_cap = dd.get('Discharge_Capacity(Ah)', np.array([]))
    if len(cycle) > 0 and len(discharge_cap) > 0:
        min_len = min(len(cycle), len(discharge_cap))
        ax.plot(cycle[:min_len], discharge_cap[:min_len], 'o-', markersize=3, label=key)

ax.set_xlabel('Cycle Index')
ax.set_ylabel('Discharge Capacity (Ah)')
ax.legend()
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'cs2_36_capacity_fade.png'))
plt.close()
print("Saved cs2_36_capacity_fade.png")

# ============================================================
# 2. NASA PCoE Dataset
# ============================================================
print("\nLoading NASA PCoE dataset...")
nasa_dir = os.path.join(DATA_DIR, 'NASA PCoE Dataset Repository', '1. BatteryAgingARC-FY08Q4')
nasa_files = sorted([f for f in os.listdir(nasa_dir) if f.endswith('.mat')])

nasa_discharge_caps = {}
nasa_sample_voltages = {}
nasa_sample_temps = {}

for f in nasa_files:
    bname = f.replace('.mat', '')
    mat = sio.loadmat(os.path.join(nasa_dir, f))

    # Navigate: mat[bname] -> 'cycle' -> array of structs
    raw = mat[bname]
    # raw is (1,1) structured array with field 'cycle'
    cycles = raw['cycle'][0, 0]  # This gives the cycle array

    caps = []
    voltages = []
    temps = []

    for i in range(cycles.shape[1]):
        cycle = cycles[0, i]
        ctype = str(cycle['type'][0]).strip()
        if ctype == 'discharge':
            data = cycle['data'][0, 0]
            if 'Capacity' in data.dtype.names:
                cap = data['Capacity'].flatten()
                if len(cap) > 0:
                    caps.append(float(cap[-1]))
            if len(voltages) < 5:
                v = data['Voltage_measured'].flatten()
                t = data['Time'].flatten()
                temp = data['Temperature_measured'].flatten() if 'Temperature_measured' in data.dtype.names else None
                voltages.append((t - t[0], v))
                if temp is not None:
                    temps.append((t - t[0], temp))

    nasa_discharge_caps[bname] = np.array(caps)
    nasa_sample_voltages[bname] = voltages
    nasa_sample_temps[bname] = temps
    print(f"  {bname}: {len(caps)} discharge cycles, capacity range [{caps[0]:.3f}, {caps[-1]:.3f}] Ah")

# Plot NASA capacity fade
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('NASA PCoE - Battery Capacity Fade', fontsize=14, fontweight='bold')
for idx, (bname, caps) in enumerate(nasa_discharge_caps.items()):
    ax = axes[idx // 2][idx % 2]
    ax.plot(range(len(caps)), caps, 'ro-', markersize=2, linewidth=1)
    ax.set_xlabel('Discharge Cycle')
    ax.set_ylabel('Capacity (Ah)')
    ax.set_title(f'{bname} Capacity Fade')
    ax.grid(True, alpha=0.3)
    ax.axhline(y=1.4, color='k', linestyle='--', alpha=0.5, label='EOL (1.4 Ah)')
    ax.legend()
plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'nasa_capacity_fade.png'))
plt.close()
print("Saved nasa_capacity_fade.png")

# Plot NASA sample discharge curves
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle('NASA B0005 - Sample Discharge Curves', fontsize=14, fontweight='bold')
for i, (t, v) in enumerate(nasa_sample_voltages.get('B0005', [])[:5]):
    axes[0].plot(t, v, linewidth=1, label=f'Cycle {i+1}')
for i, (t, temp) in enumerate(nasa_sample_temps.get('B0005', [])[:5]):
    axes[1].plot(t, temp, linewidth=1, label=f'Cycle {i+1}')
axes[0].set_xlabel('Time (s)'); axes[0].set_ylabel('Voltage (V)')
axes[0].set_title('Discharge Voltage'); axes[0].legend(fontsize=8); axes[0].grid(True, alpha=0.3)
axes[1].set_xlabel('Time (s)'); axes[1].set_ylabel('Temperature (°C)')
axes[1].set_title('Temperature Profile'); axes[1].legend(fontsize=8); axes[1].grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'nasa_discharge_curves.png'))
plt.close()
print("Saved nasa_discharge_curves.png")

# ============================================================
# 3. Oxford Battery Degradation Dataset
# ============================================================
print("\nLoading Oxford Battery Degradation Dataset...")
oxford_dir = os.path.join(DATA_DIR, 'Oxford Battery Degradation Dataset')

# Load ExampleDC_C1.mat
oxford_example = sio.loadmat(os.path.join(oxford_dir, 'ExampleDC_C1.mat'), squeeze_me=True)

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('Oxford Battery - Example Drive Cycle', fontsize=14, fontweight='bold')

for key in oxford_example.keys():
    if key.startswith('__'):
        continue
    obj = oxford_example[key]
    print(f"  Key: {key}, type={type(obj)}")
    if hasattr(obj, 'dtype') and obj.dtype.names:
        print(f"    Fields: {obj.dtype.names}")
        for name in obj.dtype.names:
            sub = obj[name].item() if hasattr(obj[name], 'item') else obj[name]
            if hasattr(sub, 'dtype') and sub.dtype.names:
                print(f"      {name}: fields={sub.dtype.names}")

                def extract_field(s, field):
                    x = s[field]
                    if x.ndim == 0:
                        x = x.item()
                    return np.array(x, dtype=float).flatten()

                t_data = extract_field(sub, 't') if 't' in sub.dtype.names else None
                v_data = extract_field(sub, 'v') if 'v' in sub.dtype.names else None
                i_data = extract_field(sub, 'i') if 'i' in sub.dtype.names else None

                if t_data is not None and v_data is not None:
                    ax_idx = 0 if 'ch' in name.lower() else 1
                    axes[0][ax_idx].plot(t_data, v_data, 'b-', linewidth=0.5)
                    axes[0][ax_idx].set_xlabel('Time (s)')
                    axes[0][ax_idx].set_ylabel('Voltage (V)')
                    axes[0][ax_idx].set_title(f'{name} - Voltage')
                    axes[0][ax_idx].grid(True, alpha=0.3)

                    if i_data is not None:
                        axes[1][ax_idx].plot(t_data, i_data, 'r-', linewidth=0.5)
                        axes[1][ax_idx].set_xlabel('Time (s)')
                        axes[1][ax_idx].set_ylabel('Current (mA)')
                        axes[1][ax_idx].set_title(f'{name} - Current')
                        axes[1][ax_idx].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'oxford_drive_cycle.png'))
plt.close()
print("Saved oxford_drive_cycle.png")

# Save extracted NASA data for later use
np.savez(os.path.join(OUT_DIR, 'nasa_extracted.npz'),
         **{f'{k}_caps': v for k, v in nasa_discharge_caps.items()})

print("\nData exploration complete!")
