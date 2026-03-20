"""
03_mmga_framework.py - Meta-Model based Genetic Algorithm (MMGA) for battery
parameter identification.

Pipeline:
1. Latin Hypercube Sampling (LHS) to generate parameter samples
2. Run SPM simulations to generate training data
3. Train ANN meta-model to predict voltage/temperature curves from parameters
4. Use GA to optimize parameters against experimental data using ANN as surrogate
"""
import os, sys, time, json
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pyDOE2 import lhs
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import warnings
warnings.filterwarnings('ignore')

# Add code directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from spm_model import SPMSimulator, SPMParameters, PARAM_BOUNDS, PARAM_NAMES, N_PARAMS
from spm_model import params_from_vector, vector_from_params

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG_DIR = os.path.join(WORKSPACE, 'report', 'images')
OUT_DIR = os.path.join(WORKSPACE, 'outputs')
os.makedirs(IMG_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# Step 1: Latin Hypercube Sampling
# ============================================================
def generate_lhs_samples(n_samples=500, seed=42):
    """Generate LHS samples in [0,1]^d normalized parameter space."""
    np.random.seed(seed)
    samples = lhs(N_PARAMS, samples=n_samples, criterion='maximin')
    return samples  # shape (n_samples, N_PARAMS)


# ============================================================
# Step 2: Run SPM simulations for training data
# ============================================================
def run_simulations(samples, I_app=2.0, t_final=4500, dt=2.0, n_output_pts=100):
    """
    Run SPM for each parameter sample. Returns feature matrix.

    For each sample, extract a fixed-length feature vector from the discharge curve:
    - Voltage at n_output_pts equally-spaced capacity points
    - Final temperature
    - Total capacity
    """
    n_samples = len(samples)
    # Output: voltage at fixed capacity fractions + temp + total_cap
    n_features = n_output_pts + 2
    Y = np.zeros((n_samples, n_features))
    valid_mask = np.ones(n_samples, dtype=bool)

    cap_fracs = np.linspace(0, 1, n_output_pts)

    for i in range(n_samples):
        if i % 50 == 0:
            print(f"  Simulating sample {i}/{n_samples}...")
        try:
            p = params_from_vector(samples[i])
            sim = SPMSimulator(p)
            result = sim.simulate_discharge(I_app, t_final, dt=dt, V_cutoff=2.5, n_r=8)

            v = result['voltage']
            cap = result['capacity']
            temp = result['temperature']

            if len(v) < 10 or cap[-1] < 0.1:
                valid_mask[i] = False
                continue

            # Interpolate voltage at fixed capacity fractions
            cap_norm = cap / (cap[-1] + 1e-10)
            interp_func = interp1d(cap_norm, v, kind='linear', fill_value='extrapolate')
            v_interp = interp_func(cap_fracs)

            Y[i, :n_output_pts] = v_interp
            Y[i, n_output_pts] = temp[-1]  # Final temperature
            Y[i, n_output_pts + 1] = cap[-1]  # Total capacity

        except Exception as e:
            valid_mask[i] = False

    print(f"  Valid simulations: {np.sum(valid_mask)}/{n_samples}")
    return Y, valid_mask

from scipy.interpolate import interp1d


# ============================================================
# Step 3: Train ANN Meta-Model
# ============================================================
class ANNMetaModel:
    """ANN surrogate model that predicts discharge features from parameters."""

    def __init__(self):
        self.scaler_X = StandardScaler()
        self.scaler_Y = StandardScaler()
        self.model = None
        self.train_score = None
        self.test_score = None

    def train(self, X, Y, test_size=0.2, seed=42):
        """Train ANN on (parameters -> discharge features) mapping."""
        X_train, X_test, Y_train, Y_test = train_test_split(
            X, Y, test_size=test_size, random_state=seed)

        X_train_s = self.scaler_X.fit_transform(X_train)
        X_test_s = self.scaler_X.transform(X_test)
        Y_train_s = self.scaler_Y.fit_transform(Y_train)
        Y_test_s = self.scaler_Y.transform(Y_test)

        self.model = MLPRegressor(
            hidden_layer_sizes=(128, 64, 32),
            activation='relu',
            solver='adam',
            max_iter=1000,
            early_stopping=True,
            validation_fraction=0.15,
            n_iter_no_change=20,
            random_state=seed,
            batch_size=32,
            learning_rate='adaptive',
            learning_rate_init=0.001,
        )

        print("  Training ANN meta-model...")
        t0 = time.time()
        self.model.fit(X_train_s, Y_train_s)
        train_time = time.time() - t0

        self.train_score = self.model.score(X_train_s, Y_train_s)
        self.test_score = self.model.score(X_test_s, Y_test_s)

        print(f"  Training time: {train_time:.1f}s")
        print(f"  Train R²: {self.train_score:.4f}")
        print(f"  Test R²: {self.test_score:.4f}")

        return self.train_score, self.test_score

    def predict(self, X):
        """Predict discharge features from parameter vectors."""
        X_s = self.scaler_X.transform(X.reshape(-1, N_PARAMS))
        Y_s = self.model.predict(X_s)
        Y = self.scaler_Y.inverse_transform(Y_s)
        return Y


# ============================================================
# Step 4: Genetic Algorithm Optimization
# ============================================================
class GeneticAlgorithm:
    """GA optimizer using ANN meta-model as fitness function."""

    def __init__(self, meta_model, target_features, pop_size=100, n_gen=200,
                 crossover_rate=0.8, mutation_rate=0.1, elite_frac=0.1, seed=42):
        self.meta = meta_model
        self.target = target_features  # shape (n_features,)
        self.pop_size = pop_size
        self.n_gen = n_gen
        self.cr = crossover_rate
        self.mr = mutation_rate
        self.elite_n = max(2, int(pop_size * elite_frac))
        self.rng = np.random.RandomState(seed)

    def fitness(self, individual):
        """Compute fitness (negative MSE between predicted and target features)."""
        pred = self.meta.predict(individual).flatten()
        # Weighted MSE: more weight on voltage, less on temperature
        n_v = len(pred) - 2
        mse_v = np.mean((pred[:n_v] - self.target[:n_v])**2)
        mse_t = (pred[n_v] - self.target[n_v])**2
        mse_cap = (pred[n_v+1] - self.target[n_v+1])**2
        return -(mse_v + 0.01 * mse_t + 0.1 * mse_cap)

    def run(self):
        """Run GA optimization."""
        # Initialize population
        pop = self.rng.rand(self.pop_size, N_PARAMS)

        best_fitness_history = []
        mean_fitness_history = []
        best_individual = None
        best_fit = -np.inf

        for gen in range(self.n_gen):
            # Evaluate fitness
            fits = np.array([self.fitness(ind) for ind in pop])

            # Track best
            gen_best_idx = np.argmax(fits)
            if fits[gen_best_idx] > best_fit:
                best_fit = fits[gen_best_idx]
                best_individual = pop[gen_best_idx].copy()

            best_fitness_history.append(best_fit)
            mean_fitness_history.append(np.mean(fits))

            if gen % 50 == 0:
                print(f"  Gen {gen}: best_fit={best_fit:.6f}, mean_fit={np.mean(fits):.6f}")

            # Selection (tournament)
            new_pop = []

            # Elitism
            elite_idx = np.argsort(fits)[-self.elite_n:]
            for idx in elite_idx:
                new_pop.append(pop[idx].copy())

            # Fill rest with crossover + mutation
            while len(new_pop) < self.pop_size:
                # Tournament selection
                t_size = 3
                idx1 = self.rng.choice(self.pop_size, t_size)
                p1 = pop[idx1[np.argmax(fits[idx1])]]
                idx2 = self.rng.choice(self.pop_size, t_size)
                p2 = pop[idx2[np.argmax(fits[idx2])]]

                # Crossover (SBX)
                if self.rng.rand() < self.cr:
                    mask = self.rng.rand(N_PARAMS) < 0.5
                    child = np.where(mask, p1, p2)
                else:
                    child = p1.copy()

                # Mutation (Gaussian)
                if self.rng.rand() < self.mr:
                    n_mut = self.rng.randint(1, max(2, N_PARAMS // 3))
                    mut_idx = self.rng.choice(N_PARAMS, n_mut, replace=False)
                    child[mut_idx] += self.rng.randn(n_mut) * 0.1
                    child = np.clip(child, 0, 1)

                new_pop.append(child)

            pop = np.array(new_pop[:self.pop_size])

        return best_individual, best_fit, best_fitness_history, mean_fitness_history


# ============================================================
# Step 5: Extract experimental target features
# ============================================================
def extract_cs2_target(n_output_pts=100):
    """Extract target features from CS2_36 dataset."""
    import openpyxl
    cs2_dir = os.path.join(WORKSPACE, 'data', 'CS2_36')
    # Use first file, first discharge cycle
    fpath = os.path.join(cs2_dir, 'CS2_36_1_10_11.xlsx')
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if 'Channel' not in sheet_name:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else '' for h in rows[0]]

        data = {}
        for i, h in enumerate(headers):
            vals = []
            for r in rows[1:]:
                if r[i] is not None:
                    try:
                        vals.append(float(r[i]))
                    except (ValueError, TypeError):
                        pass
            data[h] = np.array(vals)

        voltage = data.get('Voltage(V)', np.array([]))
        current = data.get('Current(A)', np.array([]))
        test_time = data.get('Test_Time(s)', np.array([]))

        if len(voltage) == 0:
            continue

        # Find first discharge cycle (negative current)
        discharge_mask = current < -0.01
        if np.sum(discharge_mask) < 10:
            continue

        # Get contiguous discharge segments
        diff = np.diff(discharge_mask.astype(int))
        starts = np.where(diff == 1)[0] + 1
        ends = np.where(diff == -1)[0] + 1

        if len(starts) == 0:
            if discharge_mask[0]:
                starts = np.array([0])
            else:
                continue
        if len(ends) == 0 or (len(ends) > 0 and ends[-1] < starts[-1]):
            ends = np.append(ends, len(discharge_mask))

        # Use first complete discharge
        s, e = starts[0], ends[0]
        v_dis = voltage[s:e]
        t_dis = test_time[s:e] - test_time[s]
        i_dis = np.abs(current[s:e])

        # Compute capacity (Ah)
        dt_arr = np.diff(t_dis)
        cap = np.cumsum(np.append(0, i_dis[1:] * dt_arr / 3600.0))

        # Interpolate to fixed capacity fractions
        cap_norm = cap / (cap[-1] + 1e-10)
        cap_fracs = np.linspace(0, 1, n_output_pts)
        v_interp = interp1d(cap_norm, v_dis, kind='linear', fill_value='extrapolate')(cap_fracs)

        # Target features: interpolated voltages + assumed final temp + total capacity
        target = np.zeros(n_output_pts + 2)
        target[:n_output_pts] = v_interp
        target[n_output_pts] = 30.0  # Approximate room temperature rise
        target[n_output_pts + 1] = cap[-1]

        wb.close()
        print(f"  CS2_36 target: capacity={cap[-1]:.3f} Ah, V_range=[{v_dis[-1]:.3f}, {v_dis[0]:.3f}]")
        return target, v_dis, t_dis, cap

    wb.close()
    return None, None, None, None


def extract_nasa_target(battery='B0005', cycle_idx=0, n_output_pts=100):
    """Extract target features from NASA dataset."""
    import scipy.io as sio
    nasa_dir = os.path.join(WORKSPACE, 'data', 'NASA PCoE Dataset Repository',
                            '1. BatteryAgingARC-FY08Q4')
    mat = sio.loadmat(os.path.join(nasa_dir, f'{battery}.mat'))
    cycles = mat[battery]['cycle'][0, 0]

    discharge_count = 0
    for i in range(cycles.shape[1]):
        cycle = cycles[0, i]
        ctype = str(cycle['type'][0]).strip()
        if ctype == 'discharge':
            if discharge_count == cycle_idx:
                data = cycle['data'][0, 0]
                v = data['Voltage_measured'].flatten()
                t = data['Time'].flatten() - data['Time'].flatten()[0]
                i_meas = np.abs(data['Current_measured'].flatten())
                temp = data['Temperature_measured'].flatten()

                # Compute capacity
                dt_arr = np.diff(t)
                cap = np.cumsum(np.append(0, i_meas[1:] * dt_arr / 3600.0))

                cap_norm = cap / (cap[-1] + 1e-10)
                cap_fracs = np.linspace(0, 1, n_output_pts)
                v_interp = interp1d(cap_norm, v, kind='linear', fill_value='extrapolate')(cap_fracs)

                target = np.zeros(n_output_pts + 2)
                target[:n_output_pts] = v_interp
                target[n_output_pts] = temp[-1]
                target[n_output_pts + 1] = cap[-1]

                print(f"  NASA {battery} cycle {cycle_idx}: cap={cap[-1]:.3f} Ah, "
                      f"V=[{v[-1]:.3f}, {v[0]:.3f}], T_final={temp[-1]:.1f}°C")
                return target, v, t, cap, temp
            discharge_count += 1

    return None, None, None, None, None


# ============================================================
# Main Pipeline
# ============================================================
if __name__ == '__main__':
    np.random.seed(42)

    N_SAMPLES = 300
    N_OUTPUT_PTS = 50
    POP_SIZE = 80
    N_GEN = 150

    print("=" * 60)
    print("MMGA Framework for Battery Parameter Identification")
    print("=" * 60)

    # --- Step 1: LHS Sampling ---
    print("\n[Step 1] Generating LHS samples...")
    samples = generate_lhs_samples(N_SAMPLES)
    print(f"  Generated {N_SAMPLES} samples in {N_PARAMS}D space")

    # Visualize LHS
    fig, axes = plt.subplots(2, 3, figsize=(15, 9))
    fig.suptitle('Latin Hypercube Sampling - Parameter Distributions', fontweight='bold')
    for i, ax in enumerate(axes.flatten()):
        if i < N_PARAMS:
            ax.hist(samples[:, i], bins=30, alpha=0.7, edgecolor='black', linewidth=0.5)
            name = PARAM_NAMES[i]
            lb, ub, log_scale, desc = PARAM_BOUNDS[name]
            ax.set_title(name, fontsize=9)
            ax.set_xlabel('Normalized value')
        else:
            ax.axis('off')
    plt.tight_layout()
    plt.savefig(os.path.join(IMG_DIR, 'lhs_distributions.png'))
    plt.close()
    print("  Saved lhs_distributions.png")

    # --- Step 2: Run SPM simulations ---
    print("\n[Step 2] Running SPM simulations for training data...")
    t0 = time.time()
    Y, valid_mask = run_simulations(samples, I_app=2.0, t_final=4500, dt=2.0,
                                    n_output_pts=N_OUTPUT_PTS)
    sim_time = time.time() - t0
    print(f"  Simulation time: {sim_time:.1f}s")

    X_valid = samples[valid_mask]
    Y_valid = Y[valid_mask]
    print(f"  Training data: {len(X_valid)} valid samples")

    # Save training data
    np.savez(os.path.join(OUT_DIR, 'training_data.npz'),
             X=X_valid, Y=Y_valid, param_names=PARAM_NAMES)

    # --- Step 3: Train ANN Meta-Model ---
    print("\n[Step 3] Training ANN Meta-Model...")
    meta = ANNMetaModel()
    train_r2, test_r2 = meta.train(X_valid, Y_valid)

    # Plot training loss
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(meta.model.loss_curve_, 'b-', linewidth=1)
    ax.set_xlabel('Iteration')
    ax.set_ylabel('Loss')
    ax.set_title(f'ANN Training Loss (Train R²={train_r2:.4f}, Test R²={test_r2:.4f})')
    ax.set_yscale('log')
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(IMG_DIR, 'ann_training_loss.png'))
    plt.close()
    print("  Saved ann_training_loss.png")

    # --- Step 4a: Extract CS2_36 target ---
    print("\n[Step 4a] Extracting CS2_36 experimental target...")
    cs2_target, cs2_v, cs2_t, cs2_cap = extract_cs2_target(N_OUTPUT_PTS)

    # --- Step 4b: GA Optimization for CS2_36 ---
    print("\n[Step 4b] Running GA optimization for CS2_36...")
    t0 = time.time()
    ga = GeneticAlgorithm(meta, cs2_target, pop_size=POP_SIZE, n_gen=N_GEN)
    best_vec, best_fit, best_hist, mean_hist = ga.run()
    ga_time = time.time() - t0
    print(f"  GA optimization time: {ga_time:.1f}s")
    print(f"  Best fitness: {best_fit:.6f}")

    # Convert best vector to parameters
    best_params = params_from_vector(best_vec)
    identified_params = {}
    print("\n  Identified Parameters:")
    for i, name in enumerate(PARAM_NAMES):
        val = getattr(best_params, name)
        lb, ub, log_scale, desc = PARAM_BOUNDS[name]
        identified_params[name] = val
        print(f"    {name}: {val:.6e}  [{lb:.2e}, {ub:.2e}]  ({desc})")

    # Save identified parameters
    with open(os.path.join(OUT_DIR, 'identified_params_cs2.json'), 'w') as f:
        json.dump(identified_params, f, indent=2)

    # --- Step 5: Validate on CS2_36 ---
    print("\n[Step 5] Validating identified parameters on CS2_36...")
    sim_best = SPMSimulator(best_params)
    result_best = sim_best.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.5, n_r=10)

    # Plot GA convergence
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(best_hist, 'b-', label='Best fitness', linewidth=1.5)
    ax.plot(mean_hist, 'r--', label='Mean fitness', linewidth=1, alpha=0.7)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Fitness (negative MSE)')
    ax.set_title('GA Convergence')
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(IMG_DIR, 'ga_convergence.png'))
    plt.close()
    print("  Saved ga_convergence.png")

    # Plot validation: simulated vs experimental
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    fig.suptitle('CS2_36 Validation: Identified Model vs Experimental', fontweight='bold')

    if cs2_cap is not None:
        axes[0].plot(cs2_cap, cs2_v, 'k-', linewidth=2, label='Experimental')
    axes[0].plot(result_best['capacity'], result_best['voltage'], 'r--', linewidth=2,
                label='Identified SPM')
    axes[0].set_xlabel('Capacity (Ah)')
    axes[0].set_ylabel('Voltage (V)')
    axes[0].set_title('Voltage vs Capacity')
    axes[0].legend()
    axes[0].grid(True, alpha=0.3)

    if cs2_t is not None:
        axes[1].plot(cs2_t, cs2_v, 'k-', linewidth=2, label='Experimental')
    axes[1].plot(result_best['time'], result_best['voltage'], 'r--', linewidth=2,
                label='Identified SPM')
    axes[1].set_xlabel('Time (s)')
    axes[1].set_ylabel('Voltage (V)')
    axes[1].set_title('Voltage vs Time')
    axes[1].legend()
    axes[1].grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(os.path.join(IMG_DIR, 'cs2_validation.png'))
    plt.close()
    print("  Saved cs2_validation.png")

    # --- Step 6: Validate on NASA dataset ---
    print("\n[Step 6] Validating on NASA B0005...")
    nasa_target, nasa_v, nasa_t, nasa_cap, nasa_temp = extract_nasa_target(
        'B0005', cycle_idx=0, n_output_pts=N_OUTPUT_PTS)

    # Run GA for NASA
    print("  Running GA for NASA B0005...")
    ga_nasa = GeneticAlgorithm(meta, nasa_target, pop_size=POP_SIZE, n_gen=N_GEN)
    best_vec_nasa, best_fit_nasa, best_hist_nasa, mean_hist_nasa = ga_nasa.run()
    best_params_nasa = params_from_vector(best_vec_nasa)

    # Simulate with NASA-identified parameters
    sim_nasa = SPMSimulator(best_params_nasa)
    result_nasa = sim_nasa.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.2, n_r=10)

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle('NASA B0005 Validation: Identified Model vs Experimental', fontweight='bold')

    axes[0].plot(nasa_cap, nasa_v, 'k-', linewidth=2, label='Experimental')
    axes[0].plot(result_nasa['capacity'], result_nasa['voltage'], 'r--', linewidth=2,
                label='Identified SPM')
    axes[0].set_xlabel('Capacity (Ah)'); axes[0].set_ylabel('Voltage (V)')
    axes[0].set_title('Voltage vs Capacity'); axes[0].legend(); axes[0].grid(True, alpha=0.3)

    axes[1].plot(nasa_t, nasa_v, 'k-', linewidth=2, label='Experimental')
    axes[1].plot(result_nasa['time'], result_nasa['voltage'], 'r--', linewidth=2,
                label='Identified SPM')
    axes[1].set_xlabel('Time (s)'); axes[1].set_ylabel('Voltage (V)')
    axes[1].set_title('Voltage vs Time'); axes[1].legend(); axes[1].grid(True, alpha=0.3)

    axes[2].plot(nasa_t, nasa_temp, 'k-', linewidth=2, label='Experimental')
    axes[2].plot(result_nasa['time'], result_nasa['temperature'], 'r--', linewidth=2,
                label='Identified SPM')
    axes[2].set_xlabel('Time (s)'); axes[2].set_ylabel('Temperature (°C)')
    axes[2].set_title('Temperature'); axes[2].legend(); axes[2].grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(os.path.join(IMG_DIR, 'nasa_validation.png'))
    plt.close()
    print("  Saved nasa_validation.png")

    # Save NASA identified params
    nasa_identified = {name: float(getattr(best_params_nasa, name)) for name in PARAM_NAMES}
    with open(os.path.join(OUT_DIR, 'identified_params_nasa.json'), 'w') as f:
        json.dump(nasa_identified, f, indent=2)

    # --- Step 7: Compute error metrics ---
    print("\n[Step 7] Computing error metrics...")

    # CS2_36 error
    if cs2_cap is not None and len(result_best['voltage']) > 10:
        # Interpolate both to same capacity points
        cap_max = min(cs2_cap[-1], result_best['capacity'][-1])
        cap_eval = np.linspace(0, cap_max * 0.95, 100)

        v_exp_interp = interp1d(cs2_cap, cs2_v, fill_value='extrapolate')(cap_eval)
        v_sim_interp = interp1d(result_best['capacity'], result_best['voltage'],
                               fill_value='extrapolate')(cap_eval)

        rmse_cs2 = np.sqrt(np.mean((v_exp_interp - v_sim_interp)**2))
        mae_cs2 = np.mean(np.abs(v_exp_interp - v_sim_interp))
        print(f"  CS2_36: RMSE={rmse_cs2*1000:.1f} mV, MAE={mae_cs2*1000:.1f} mV")
    else:
        rmse_cs2 = mae_cs2 = float('nan')

    # NASA error
    if nasa_cap is not None and len(result_nasa['voltage']) > 10:
        cap_max = min(nasa_cap[-1], result_nasa['capacity'][-1])
        cap_eval = np.linspace(0, cap_max * 0.95, 100)

        v_exp_interp = interp1d(nasa_cap, nasa_v, fill_value='extrapolate')(cap_eval)
        v_sim_interp = interp1d(result_nasa['capacity'], result_nasa['voltage'],
                               fill_value='extrapolate')(cap_eval)

        rmse_nasa = np.sqrt(np.mean((v_exp_interp - v_sim_interp)**2))
        mae_nasa = np.mean(np.abs(v_exp_interp - v_sim_interp))
        print(f"  NASA B0005: RMSE={rmse_nasa*1000:.1f} mV, MAE={mae_nasa*1000:.1f} mV")
    else:
        rmse_nasa = mae_nasa = float('nan')

    # --- Step 8: Speed comparison ---
    print("\n[Step 8] Speed comparison: SPM vs ANN meta-model...")

    # Time SPM
    t0 = time.time()
    for _ in range(10):
        sim = SPMSimulator()
        sim.simulate_discharge(2.0, 4000, dt=2.0, n_r=8)
    spm_time = (time.time() - t0) / 10

    # Time ANN
    test_vecs = np.random.rand(1000, N_PARAMS)
    t0 = time.time()
    meta.predict(test_vecs[0])  # warmup
    t0 = time.time()
    for i in range(1000):
        meta.predict(test_vecs[i])
    ann_time = (time.time() - t0) / 1000

    speedup = spm_time / ann_time
    print(f"  SPM simulation: {spm_time*1000:.1f} ms/run")
    print(f"  ANN prediction: {ann_time*1000:.3f} ms/run")
    print(f"  Speedup: {speedup:.0f}x")

    # --- Save all results ---
    results = {
        'cs2_rmse_mV': float(rmse_cs2 * 1000) if not np.isnan(rmse_cs2) else None,
        'cs2_mae_mV': float(mae_cs2 * 1000) if not np.isnan(mae_cs2) else None,
        'nasa_rmse_mV': float(rmse_nasa * 1000) if not np.isnan(rmse_nasa) else None,
        'nasa_mae_mV': float(mae_nasa * 1000) if not np.isnan(mae_nasa) else None,
        'ann_train_r2': float(train_r2),
        'ann_test_r2': float(test_r2),
        'spm_time_ms': float(spm_time * 1000),
        'ann_time_ms': float(ann_time * 1000),
        'speedup': float(speedup),
        'n_lhs_samples': N_SAMPLES,
        'n_ga_generations': N_GEN,
        'ga_population_size': POP_SIZE,
        'total_sim_time_s': float(sim_time),
        'ga_time_s': float(ga_time),
    }

    with open(os.path.join(OUT_DIR, 'mmga_results.json'), 'w') as f:
        json.dump(results, f, indent=2)

    print("\n" + "=" * 60)
    print("MMGA Framework complete!")
    print("=" * 60)
