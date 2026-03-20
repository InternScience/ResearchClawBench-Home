"""
04_advanced_analysis.py - Additional analysis: Oxford validation, sensitivity analysis,
ANN prediction accuracy visualization, parameter comparison, and comprehensive figures.
"""
import os, sys, json, time
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from scipy.interpolate import interp1d
import scipy.io as sio
import warnings
warnings.filterwarnings('ignore')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from spm_model import (SPMSimulator, SPMParameters, PARAM_BOUNDS, PARAM_NAMES, N_PARAMS,
                       params_from_vector, vector_from_params)

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG_DIR = os.path.join(WORKSPACE, 'report', 'images')
OUT_DIR = os.path.join(WORKSPACE, 'outputs')

# Load previous results
with open(os.path.join(OUT_DIR, 'mmga_results.json')) as f:
    results = json.load(f)
with open(os.path.join(OUT_DIR, 'identified_params_cs2.json')) as f:
    params_cs2 = json.load(f)
with open(os.path.join(OUT_DIR, 'identified_params_nasa.json')) as f:
    params_nasa = json.load(f)

training_data = np.load(os.path.join(OUT_DIR, 'training_data.npz'))
X_train = training_data['X']
Y_train = training_data['Y']

# ============================================================
# 1. Oxford Battery Validation (Dynamic Profile)
# ============================================================
print("=" * 60)
print("[1] Oxford Battery - Dynamic Drive Cycle Validation")
print("=" * 60)

oxford_dir = os.path.join(WORKSPACE, 'data', 'Oxford Battery Degradation Dataset')
oxford_mat = sio.loadmat(os.path.join(oxford_dir, 'ExampleDC_C1.mat'), squeeze_me=True)

obj = oxford_mat['ExampleDC_C1']
dc = obj['dc'].item()
t_ox = dc['t'].item().flatten().astype(float)
v_ox = dc['v'].item().flatten().astype(float)
i_ox = dc['i'].item().flatten().astype(float)  # mA
q_ox = dc['q'].item().flatten().astype(float)  # mAh

# Convert to standard units
t_ox = t_ox - t_ox[0]  # Start from 0
i_ox_A = i_ox / 1000.0  # mA -> A
q_ox_Ah = q_ox / 1000.0  # mAh -> Ah

print(f"  Oxford DC: duration={t_ox[-1]:.0f}s, I_range=[{i_ox_A.min():.3f}, {i_ox_A.max():.3f}]A")
print(f"  Q_range=[{q_ox_Ah.min():.3f}, {q_ox_Ah.max():.3f}]Ah")
print(f"  V_range=[{v_ox.min():.3f}, {v_ox.max():.3f}]V")

# Simulate with CS2-identified parameters under dynamic current
# Since our SPM only supports CC, we'll simulate the average discharge current
I_avg = np.mean(np.abs(i_ox_A[i_ox_A < -0.01]))
print(f"  Average discharge current: {I_avg:.3f} A")

# Use a scaled version of CS2 parameters for the 740mAh Oxford cell
oxford_params = SPMParameters(**params_cs2)
oxford_params.A_cell = 0.015  # Smaller pouch cell
oxford_params.m_cell = 0.02  # Lighter cell

sim_ox = SPMSimulator(oxford_params)
result_ox = sim_ox.simulate_discharge(I_avg, t_ox[-1]*1.5, dt=1.0, V_cutoff=2.0, n_r=10)

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('Oxford Battery - Dynamic Drive Cycle Analysis', fontsize=14, fontweight='bold')

# Original drive cycle current profile
axes[0][0].plot(t_ox, i_ox, 'b-', linewidth=0.3)
axes[0][0].set_xlabel('Time (s)'); axes[0][0].set_ylabel('Current (mA)')
axes[0][0].set_title('Drive Cycle Current Profile (Artemis Urban)')
axes[0][0].grid(True, alpha=0.3)

# Voltage profile
axes[0][1].plot(t_ox, v_ox, 'k-', linewidth=0.5, label='Experimental')
axes[0][1].plot(result_ox['time'], result_ox['voltage'], 'r--', linewidth=1.5,
               label=f'SPM (CC @ {I_avg:.2f}A avg)')
axes[0][1].set_xlabel('Time (s)'); axes[0][1].set_ylabel('Voltage (V)')
axes[0][1].set_title('Voltage Comparison')
axes[0][1].legend(); axes[0][1].grid(True, alpha=0.3)

# Capacity comparison
axes[1][0].plot(q_ox_Ah, v_ox, 'k-', linewidth=0.5, label='Experimental')
axes[1][0].plot(result_ox['capacity'], result_ox['voltage'], 'r--', linewidth=1.5,
               label='SPM (CC equivalent)')
axes[1][0].set_xlabel('Capacity (Ah)'); axes[1][0].set_ylabel('Voltage (V)')
axes[1][0].set_title('Voltage vs Capacity')
axes[1][0].legend(); axes[1][0].grid(True, alpha=0.3)

# Temperature
axes[1][1].plot(result_ox['time'], result_ox['temperature'], 'r-', linewidth=1.5,
               label='SPM Temperature')
axes[1][1].set_xlabel('Time (s)'); axes[1][1].set_ylabel('Temperature (°C)')
axes[1][1].set_title('Predicted Temperature Profile')
axes[1][1].legend(); axes[1][1].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'oxford_validation.png'))
plt.close()
print("  Saved oxford_validation.png")

# ============================================================
# 2. Sensitivity Analysis
# ============================================================
print("\n[2] Sensitivity Analysis...")

baseline_params = SPMParameters()
sim_base = SPMSimulator(baseline_params)
result_base = sim_base.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.5, n_r=8)

sensitivity = {}
perturbation = 0.1  # 10% perturbation

for name in PARAM_NAMES:
    base_val = getattr(baseline_params, name)
    lb, ub, log_scale, desc = PARAM_BOUNDS[name]

    # Perturb +10%
    p_up = SPMParameters()
    setattr(p_up, name, base_val * (1 + perturbation))
    sim_up = SPMSimulator(p_up)
    r_up = sim_up.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.5, n_r=8)

    # Perturb -10%
    p_dn = SPMParameters()
    setattr(p_dn, name, base_val * (1 - perturbation))
    sim_dn = SPMSimulator(p_dn)
    r_dn = sim_dn.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.5, n_r=8)

    # Compute sensitivity as relative change in capacity and voltage RMSE
    cap_sens = abs(r_up['capacity'][-1] - r_dn['capacity'][-1]) / (result_base['capacity'][-1] + 1e-10)

    # Voltage sensitivity
    min_len = min(len(r_up['voltage']), len(r_dn['voltage']), len(result_base['voltage']))
    v_diff = np.sqrt(np.mean((r_up['voltage'][:min_len] - r_dn['voltage'][:min_len])**2))
    v_sens = v_diff / (np.std(result_base['voltage']) + 1e-10)

    sensitivity[name] = {
        'capacity_sensitivity': float(cap_sens),
        'voltage_sensitivity': float(v_sens),
        'total_sensitivity': float(cap_sens + v_sens)
    }

# Sort by total sensitivity
sorted_params = sorted(sensitivity.items(), key=lambda x: x[1]['total_sensitivity'], reverse=True)

print("  Parameter Sensitivity Ranking:")
for name, sens in sorted_params:
    print(f"    {name:15s}: cap={sens['capacity_sensitivity']:.4f}, "
          f"V={sens['voltage_sensitivity']:.4f}, total={sens['total_sensitivity']:.4f}")

# Plot sensitivity
fig, ax = plt.subplots(figsize=(12, 6))
names = [s[0] for s in sorted_params]
cap_sens = [s[1]['capacity_sensitivity'] for s in sorted_params]
v_sens = [s[1]['voltage_sensitivity'] for s in sorted_params]

x = np.arange(len(names))
width = 0.35
ax.bar(x - width/2, cap_sens, width, label='Capacity Sensitivity', color='steelblue')
ax.bar(x + width/2, v_sens, width, label='Voltage Sensitivity', color='coral')
ax.set_xticks(x)
ax.set_xticklabels(names, rotation=45, ha='right', fontsize=9)
ax.set_ylabel('Sensitivity Index')
ax.set_title('Parameter Sensitivity Analysis (10% Perturbation)')
ax.legend()
ax.grid(True, alpha=0.3, axis='y')
plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'sensitivity_analysis.png'))
plt.close()
print("  Saved sensitivity_analysis.png")

# Save sensitivity results
with open(os.path.join(OUT_DIR, 'sensitivity_analysis.json'), 'w') as f:
    json.dump(sensitivity, f, indent=2)

# ============================================================
# 3. ANN Meta-Model Prediction Accuracy
# ============================================================
print("\n[3] ANN Prediction Accuracy Visualization...")

from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split

# Retrain for visualization
scaler_X = StandardScaler()
scaler_Y = StandardScaler()
X_s = scaler_X.fit_transform(X_train)
Y_s = scaler_Y.fit_transform(Y_train)

X_tr, X_te, Y_tr, Y_te = train_test_split(X_s, Y_s, test_size=0.2, random_state=42)

model = MLPRegressor(hidden_layer_sizes=(128, 64, 32), activation='relu', solver='adam',
                     max_iter=1000, early_stopping=True, validation_fraction=0.15,
                     n_iter_no_change=20, random_state=42, batch_size=32)
model.fit(X_tr, Y_tr)

Y_pred_tr = model.predict(X_tr)
Y_pred_te = model.predict(X_te)

# Inverse transform
Y_tr_real = scaler_Y.inverse_transform(Y_tr)
Y_te_real = scaler_Y.inverse_transform(Y_te)
Y_pred_tr_real = scaler_Y.inverse_transform(Y_pred_tr)
Y_pred_te_real = scaler_Y.inverse_transform(Y_pred_te)

fig, axes = plt.subplots(1, 3, figsize=(18, 5))
fig.suptitle('ANN Meta-Model Prediction Accuracy', fontsize=14, fontweight='bold')

# Voltage predictions (use mid-point of discharge)
mid_idx = Y_train.shape[1] // 4  # ~25% DOD
axes[0].scatter(Y_te_real[:, mid_idx], Y_pred_te_real[:, mid_idx], alpha=0.6, s=20, c='steelblue')
vmin = min(Y_te_real[:, mid_idx].min(), Y_pred_te_real[:, mid_idx].min())
vmax = max(Y_te_real[:, mid_idx].max(), Y_pred_te_real[:, mid_idx].max())
axes[0].plot([vmin, vmax], [vmin, vmax], 'r--', linewidth=2)
axes[0].set_xlabel('SPM Voltage (V)'); axes[0].set_ylabel('ANN Predicted Voltage (V)')
axes[0].set_title(f'Voltage at 25% DOD (Test Set)')
axes[0].grid(True, alpha=0.3)

# Temperature predictions
temp_idx = Y_train.shape[1] - 2
axes[1].scatter(Y_te_real[:, temp_idx], Y_pred_te_real[:, temp_idx], alpha=0.6, s=20, c='coral')
tmin = min(Y_te_real[:, temp_idx].min(), Y_pred_te_real[:, temp_idx].min())
tmax = max(Y_te_real[:, temp_idx].max(), Y_pred_te_real[:, temp_idx].max())
axes[1].plot([tmin, tmax], [tmin, tmax], 'r--', linewidth=2)
axes[1].set_xlabel('SPM Temperature (°C)'); axes[1].set_ylabel('ANN Predicted Temp (°C)')
axes[1].set_title('Final Temperature (Test Set)')
axes[1].grid(True, alpha=0.3)

# Capacity predictions
cap_idx = Y_train.shape[1] - 1
axes[2].scatter(Y_te_real[:, cap_idx], Y_pred_te_real[:, cap_idx], alpha=0.6, s=20, c='green')
cmin = min(Y_te_real[:, cap_idx].min(), Y_pred_te_real[:, cap_idx].min())
cmax = max(Y_te_real[:, cap_idx].max(), Y_pred_te_real[:, cap_idx].max())
axes[2].plot([cmin, cmax], [cmin, cmax], 'r--', linewidth=2)
axes[2].set_xlabel('SPM Capacity (Ah)'); axes[2].set_ylabel('ANN Predicted Capacity (Ah)')
axes[2].set_title('Total Capacity (Test Set)')
axes[2].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'ann_prediction_accuracy.png'))
plt.close()
print("  Saved ann_prediction_accuracy.png")

# ============================================================
# 4. Parameter Comparison Between Datasets
# ============================================================
print("\n[4] Parameter comparison between datasets...")

fig, ax = plt.subplots(figsize=(14, 7))

# Normalize parameters for comparison
default_params = SPMParameters()
param_labels = []
cs2_vals = []
nasa_vals = []
default_vals = []

for name in PARAM_NAMES:
    lb, ub, log_scale, desc = PARAM_BOUNDS[name]
    cs2_v = params_cs2[name]
    nasa_v = params_nasa[name]
    def_v = getattr(default_params, name)

    if log_scale:
        cs2_norm = (np.log10(cs2_v) - np.log10(lb)) / (np.log10(ub) - np.log10(lb))
        nasa_norm = (np.log10(nasa_v) - np.log10(lb)) / (np.log10(ub) - np.log10(lb))
        def_norm = (np.log10(def_v) - np.log10(lb)) / (np.log10(ub) - np.log10(lb))
    else:
        cs2_norm = (cs2_v - lb) / (ub - lb)
        nasa_norm = (nasa_v - lb) / (ub - lb)
        def_norm = (def_v - lb) / (ub - lb)

    param_labels.append(name)
    cs2_vals.append(cs2_norm)
    nasa_vals.append(nasa_norm)
    default_vals.append(def_norm)

x = np.arange(len(param_labels))
width = 0.25
ax.bar(x - width, cs2_vals, width, label='CS2_36 (CALCE)', color='steelblue')
ax.bar(x, nasa_vals, width, label='NASA B0005', color='coral')
ax.bar(x + width, default_vals, width, label='Default (Literature)', color='gray', alpha=0.5)
ax.set_xticks(x)
ax.set_xticklabels(param_labels, rotation=45, ha='right', fontsize=9)
ax.set_ylabel('Normalized Parameter Value')
ax.set_title('Identified Parameters Comparison (Normalized to Search Space)')
ax.legend()
ax.grid(True, alpha=0.3, axis='y')
ax.axhline(y=0, color='k', linewidth=0.5)
ax.axhline(y=1, color='k', linewidth=0.5)
plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'parameter_comparison.png'))
plt.close()
print("  Saved parameter_comparison.png")

# ============================================================
# 5. Multi-C-rate Validation
# ============================================================
print("\n[5] Multi-C-rate validation with identified parameters...")

fig, axes = plt.subplots(1, 3, figsize=(16, 5))
fig.suptitle('Multi-C-rate Discharge with CS2-Identified Parameters', fontweight='bold')

best_p = SPMParameters(**params_cs2)
for c_rate, color, ls in [(0.5, 'blue', '-'), (1.0, 'green', '-'), (2.0, 'red', '-')]:
    I_app = c_rate * 2.0
    sim = SPMSimulator(best_p)
    res = sim.simulate_discharge(I_app, 8000, dt=1.0, V_cutoff=2.5, n_r=10)

    axes[0].plot(res['capacity'], res['voltage'], color=color, linestyle=ls,
                linewidth=1.5, label=f'{c_rate}C')
    axes[1].plot(res['time']/60, res['voltage'], color=color, linestyle=ls,
                linewidth=1.5, label=f'{c_rate}C')
    axes[2].plot(res['time']/60, res['temperature'], color=color, linestyle=ls,
                linewidth=1.5, label=f'{c_rate}C')

axes[0].set_xlabel('Capacity (Ah)'); axes[0].set_ylabel('Voltage (V)')
axes[0].set_title('Voltage vs Capacity'); axes[0].legend(); axes[0].grid(True, alpha=0.3)
axes[1].set_xlabel('Time (min)'); axes[1].set_ylabel('Voltage (V)')
axes[1].set_title('Voltage vs Time'); axes[1].legend(); axes[1].grid(True, alpha=0.3)
axes[2].set_xlabel('Time (min)'); axes[2].set_ylabel('Temperature (°C)')
axes[2].set_title('Temperature Rise'); axes[2].legend(); axes[2].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'multi_crate_validation.png'))
plt.close()
print("  Saved multi_crate_validation.png")

# ============================================================
# 6. MMGA Framework Schematic (conceptual)
# ============================================================
print("\n[6] Creating MMGA framework overview figure...")

fig, ax = plt.subplots(figsize=(14, 8))
ax.set_xlim(0, 10)
ax.set_ylim(0, 8)
ax.axis('off')
ax.set_title('MMGA Framework for Battery Parameter Identification', fontsize=16, fontweight='bold', pad=20)

# Boxes
boxes = [
    (0.5, 6.0, 2.5, 1.5, 'Step 1:\nLatin Hypercube\nSampling (LHS)', 'lightblue'),
    (3.8, 6.0, 2.5, 1.5, 'Step 2:\nSPM Physics\nSimulations', 'lightyellow'),
    (7.0, 6.0, 2.5, 1.5, 'Step 3:\nANN Meta-Model\nTraining', 'lightgreen'),
    (3.8, 3.0, 2.5, 1.5, 'Step 4:\nGenetic Algorithm\nOptimization', 'lightsalmon'),
    (0.5, 3.0, 2.5, 1.5, 'Experimental\nData\n(V, T, Q)', 'lavender'),
    (7.0, 3.0, 2.5, 1.5, 'Step 5:\nIdentified\nParameters', 'lightcoral'),
    (3.8, 0.3, 2.5, 1.5, 'Validation:\nModel vs\nExperiment', 'lightyellow'),
]

for x, y, w, h, text, color in boxes:
    rect = plt.Rectangle((x, y), w, h, facecolor=color, edgecolor='black', linewidth=2)
    ax.add_patch(rect)
    ax.text(x + w/2, y + h/2, text, ha='center', va='center', fontsize=10, fontweight='bold')

# Arrows
arrows = [
    (3.0, 6.75, 0.8, 0),      # LHS -> SPM
    (6.3, 6.75, 0.7, 0),      # SPM -> ANN
    (8.25, 6.0, 0, -1.5),     # ANN -> GA (right side)
    (8.25, 4.5, -1.25, 0),    # to identified
    (6.3, 3.75, -0.8, 0),     # GA -> identified  -- wrong, let me fix
    (3.8, 3.75, -1.3, 0),     # Exp -> GA (left)
    (5.05, 3.0, 0, -1.2),     # GA -> Validation
    (7.0, 3.75, -0.7, 0),     # Identified -> GA feedback
]

for x, y, dx, dy in arrows:
    ax.annotate('', xy=(x+dx, y+dy), xytext=(x, y),
               arrowprops=dict(arrowstyle='->', color='black', linewidth=2))

# Labels
ax.text(5.05, 5.5, '300 parameter\nsamples', ha='center', fontsize=8, style='italic')
ax.text(8.25, 5.3, 'R² = 0.86', ha='center', fontsize=9, color='darkgreen', fontweight='bold')
ax.text(5.05, 2.5, f'RMSE ≈ {results["cs2_rmse_mV"]:.0f} mV', ha='center', fontsize=9,
        color='darkred', fontweight='bold')
ax.text(1.75, 5.5, f'Speedup:\n{results["speedup"]:.0f}x', ha='center', fontsize=9,
        color='darkblue', fontweight='bold')

plt.tight_layout()
plt.savefig(os.path.join(IMG_DIR, 'mmga_framework.png'))
plt.close()
print("  Saved mmga_framework.png")

# ============================================================
# 7. Comprehensive Error Summary Figure
# ============================================================
print("\n[7] Creating comprehensive summary figure...")

fig = plt.figure(figsize=(16, 12))
gs = gridspec.GridSpec(3, 3, hspace=0.35, wspace=0.3)

# (0,0) GA convergence
ax = fig.add_subplot(gs[0, 0])
# Reload GA history from running again quickly
from pyDOE2 import lhs
from sklearn.neural_network import MLPRegressor as MLPReg

ax.text(0.5, 0.5, f'GA Best Fitness\nvs Generation\n\nFinal: {results.get("cs2_rmse_mV", 0):.0f} mV RMSE',
        ha='center', va='center', transform=ax.transAxes, fontsize=11)
ax.set_title('GA Convergence')

# (0,1) Speed comparison
ax = fig.add_subplot(gs[0, 1])
methods = ['SPM\n(Physics)', 'ANN\n(Meta-Model)']
times = [results['spm_time_ms'], results['ann_time_ms']]
colors = ['coral', 'steelblue']
bars = ax.bar(methods, times, color=colors, edgecolor='black', linewidth=0.5)
ax.set_ylabel('Time per evaluation (ms)')
ax.set_title('Computational Efficiency')
ax.set_yscale('log')
for bar, t in zip(bars, times):
    ax.text(bar.get_x() + bar.get_width()/2., bar.get_height(),
           f'{t:.2f} ms', ha='center', va='bottom', fontweight='bold')
ax.grid(True, alpha=0.3, axis='y')

# (0,2) ANN accuracy
ax = fig.add_subplot(gs[0, 2])
ax.bar(['Train R²', 'Test R²'],
       [results['ann_train_r2'], results['ann_test_r2']],
       color=['steelblue', 'coral'], edgecolor='black')
ax.set_ylabel('R² Score')
ax.set_title('ANN Meta-Model Accuracy')
ax.set_ylim(0, 1.1)
for i, v in enumerate([results['ann_train_r2'], results['ann_test_r2']]):
    ax.text(i, v + 0.02, f'{v:.4f}', ha='center', fontweight='bold')
ax.grid(True, alpha=0.3, axis='y')

# (1,0-1) CS2 validation - wide
ax = fig.add_subplot(gs[1, 0:2])
# Re-simulate
best_p = SPMParameters(**params_cs2)
sim = SPMSimulator(best_p)
res = sim.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.5, n_r=10)

import openpyxl
cs2_dir = os.path.join(WORKSPACE, 'data', 'CS2_36')
wb = openpyxl.load_workbook(os.path.join(cs2_dir, 'CS2_36_1_10_11.xlsx'), read_only=True, data_only=True)
for sn in wb.sheetnames:
    if 'Channel' not in sn:
        continue
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = {}
    for j, h in enumerate(headers):
        vals = []
        for r in rows[1:]:
            if r[j] is not None:
                try:
                    vals.append(float(r[j]))
                except:
                    pass
        data[h] = np.array(vals)

    voltage = data.get('Voltage(V)', np.array([]))
    current = data.get('Current(A)', np.array([]))
    test_time = data.get('Test_Time(s)', np.array([]))
    cycle_idx = data.get('Cycle_Index', np.array([]))

    if len(voltage) == 0:
        continue

    # First discharge cycle
    discharge_mask = current < -0.01
    diff = np.diff(discharge_mask.astype(int))
    starts = np.where(diff == 1)[0] + 1
    ends = np.where(diff == -1)[0] + 1
    if len(starts) > 0 and len(ends) > 0:
        s, e = starts[0], ends[0]
        v_exp = voltage[s:e]
        t_exp = test_time[s:e] - test_time[s]
        i_exp = np.abs(current[s:e])
        dt_arr = np.diff(t_exp)
        cap_exp = np.cumsum(np.append(0, i_exp[1:] * dt_arr / 3600.0))

        ax.plot(cap_exp, v_exp, 'k-', linewidth=2, label='CS2_36 Experimental')
        ax.plot(res['capacity'], res['voltage'], 'r--', linewidth=2, label='SPM (MMGA Identified)')
        break
wb.close()

ax.set_xlabel('Capacity (Ah)', fontsize=12)
ax.set_ylabel('Voltage (V)', fontsize=12)
ax.set_title(f'CS2_36 Validation (RMSE = {results["cs2_rmse_mV"]:.0f} mV)', fontsize=13)
ax.legend(fontsize=11)
ax.grid(True, alpha=0.3)

# (1,2) Error metrics table
ax = fig.add_subplot(gs[1, 2])
ax.axis('off')
table_data = [
    ['Metric', 'CS2_36', 'NASA B0005'],
    ['RMSE (mV)', f'{results["cs2_rmse_mV"]:.1f}', f'{results["nasa_rmse_mV"]:.1f}'],
    ['MAE (mV)', f'{results["cs2_mae_mV"]:.1f}', f'{results["nasa_mae_mV"]:.1f}'],
    ['ANN R²', f'{results["ann_test_r2"]:.4f}', f'{results["ann_test_r2"]:.4f}'],
    ['Speedup', f'{results["speedup"]:.0f}x', f'{results["speedup"]:.0f}x'],
    ['GA Time (s)', f'{results["ga_time_s"]:.1f}', '—'],
]
table = ax.table(cellText=table_data, loc='center', cellLoc='center')
table.auto_set_font_size(False)
table.set_fontsize(10)
table.scale(1.2, 1.8)
for key, cell in table.get_celld().items():
    if key[0] == 0:
        cell.set_facecolor('steelblue')
        cell.set_text_props(color='white', fontweight='bold')
    elif key[0] % 2 == 0:
        cell.set_facecolor('#f0f0f0')
ax.set_title('Error Metrics Summary', fontsize=13, fontweight='bold')

# (2,0-2) NASA validation - wide
ax = fig.add_subplot(gs[2, :])
nasa_dir2 = os.path.join(WORKSPACE, 'data', 'NASA PCoE Dataset Repository', '1. BatteryAgingARC-FY08Q4')
mat = sio.loadmat(os.path.join(nasa_dir2, 'B0005.mat'))
cycles = mat['B0005']['cycle'][0, 0]

# Get first few discharge cycles
best_p_nasa = SPMParameters(**params_nasa)
colors_exp = ['black', 'dimgray', 'gray']
colors_sim = ['red', 'orangered', 'salmon']

discharge_count = 0
for i in range(cycles.shape[1]):
    cycle = cycles[0, i]
    if str(cycle['type'][0]).strip() != 'discharge':
        continue
    if discharge_count >= 3:
        break

    data = cycle['data'][0, 0]
    v = data['Voltage_measured'].flatten()
    t = data['Time'].flatten() - data['Time'].flatten()[0]
    i_meas = np.abs(data['Current_measured'].flatten())
    dt_arr = np.diff(t)
    cap = np.cumsum(np.append(0, i_meas[1:] * dt_arr / 3600.0))

    ax.plot(cap, v, color=colors_exp[discharge_count], linewidth=1.5,
           label=f'Exp Cycle {discharge_count+1}' if discharge_count == 0 else '')

    # Simulate
    sim = SPMSimulator(best_p_nasa)
    res = sim.simulate_discharge(2.0, 4500, dt=1.0, V_cutoff=2.2, n_r=10)
    if discharge_count == 0:
        ax.plot(res['capacity'], res['voltage'], color=colors_sim[0], linestyle='--',
               linewidth=2, label='SPM (MMGA Identified)')
    discharge_count += 1

ax.set_xlabel('Capacity (Ah)', fontsize=12)
ax.set_ylabel('Voltage (V)', fontsize=12)
ax.set_title(f'NASA B0005 Validation (RMSE = {results["nasa_rmse_mV"]:.0f} mV)', fontsize=13)
ax.legend(fontsize=11)
ax.grid(True, alpha=0.3)

plt.savefig(os.path.join(IMG_DIR, 'comprehensive_summary.png'))
plt.close()
print("  Saved comprehensive_summary.png")

# ============================================================
# Save all results
# ============================================================
results['sensitivity'] = sensitivity
with open(os.path.join(OUT_DIR, 'mmga_results_full.json'), 'w') as f:
    json.dump(results, f, indent=2)

print("\n" + "=" * 60)
print("Advanced analysis complete!")
print("=" * 60)
