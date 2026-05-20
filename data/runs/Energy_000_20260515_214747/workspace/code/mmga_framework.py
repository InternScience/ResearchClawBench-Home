"""
MMGA: Meta-Model based Genetic Algorithm for Parameter Identification
of the Electrochemical-Aging-Thermal (ECAT) Coupled Model for Li-ion Batteries.

Fixed version with corrected OCP functions and stable numerics.
"""

import numpy as np
from scipy.integrate import solve_ivp
from scipy.interpolate import interp1d
from scipy.stats import qmc
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# Physical Constants
# =============================================================================
F = 96485.3329    # Faraday constant (C/mol)
R_gas = 8.314462618   # Gas constant (J/mol/K)
T_ref = 298.15    # Reference temperature (K)

# =============================================================================
# OCP Functions (properly bounded)
# =============================================================================

def ocp_nmc(x):
    """
    Open circuit potential of NMC cathode (vs Li/Li+).
    Based on literature data for NMC(111)/graphite cells.
    x is lithium stoichiometry (0 = fully delithiated, 1 = fully lithiated)
    """
    x = np.clip(x, 0.01, 0.99)
    # Use a polynomial fit that's numerically stable
    # Typical NMC OCP ranges from ~3.5V to ~4.3V
    a0, a1, a2, a3, a4, a5 = 4.6874, -4.8161, 4.2345, -1.5721, 0.2732, -0.0184
    U = a0 + a1*x + a2*x**2 + a3*x**3 + a4*x**4 + a5*x**5
    return U

def ocp_graphite(y):
    """
    Open circuit potential of graphite anode (vs Li/Li+).
    y is lithium stoichiometry (0 = fully delithiated, 1 = fully lithiated)
    """
    y = np.clip(y, 0.001, 0.999)
    # Graphite OCP ranges from ~0.05V to ~0.8V
    # Stages visible in the OCP curve
    U = (0.7222 + 0.1387 * y + 0.029 * np.sqrt(y) -
         0.0172 / (y + 1e-6) + 0.0019 / (y**1.5 + 1e-9) +
         0.2808 * np.exp(0.9 - 15 * y) -
         0.7984 * np.exp(0.4465 * y - 0.4108))
    return np.clip(U, 0.01, 1.5)

# =============================================================================
# ECAT Model: Single Particle + Aging + Thermal
# =============================================================================

class ECATModel:
    """
    Electrochemical-Aging-Thermal (ECAT) coupled model based on the
    Single Particle (SP) approximation with SEI growth aging and
    lumped thermal dynamics.
    
    Key parameters to be identified:
    - R_p, R_n: particle radii (m)
    - k_p, k_n: reaction rate constants (m^2.5 mol^-0.5 s^-1)
    - D_s_p, D_s_n: solid-phase Li diffusivities (m^2/s)
    - eps_s_p, eps_s_n: active material volume fractions
    - L_p, L_n: electrode thicknesses (m)
    - k_sei: SEI formation rate constant
    - h: heat transfer coefficient (W/m^2/K)
    """
    
    def __init__(self, params=None):
        # Default parameters for NMC/graphite 18650 cell
        self.params = {
            # Geometric
            'R_p': 5e-6,        # cathode particle radius (m)
            'R_n': 5e-6,        # anode particle radius (m)
            'L_p': 70e-6,       # cathode thickness (m)
            'L_n': 70e-6,       # anode thickness (m)
            'L_s': 25e-6,       # separator thickness (m)
            'A': 0.0784,        # electrode surface area (m^2)
            'eps_s_p': 0.50,    # cathode active material volume fraction
            'eps_s_n': 0.55,    # anode active material volume fraction
            
            # Kinetic
            'k_p': 1e-11,       # cathode reaction rate (m^2.5 mol^-0.5 s^-1)
            'k_n': 5e-11,       # anode reaction rate
            
            # Transport
            'D_s_p': 1e-13,     # cathode solid diffusivity (m^2/s)
            'D_s_n': 3.9e-14,   # anode solid diffusivity
            
            # Concentration
            'c_s_max_p': 49000, # cathode max concentration (mol/m^3)
            'c_s_max_n': 31500, # anode max concentration
            'c_e': 1000,        # electrolyte concentration (mol/m^3)
            'x0': 0.40,         # cathode stoichiometry at 100% SOC
            'y0': 0.80,         # anode stoichiometry at 100% SOC
            
            # Aging (SEI)
            'k_sei': 1e-13,     # SEI formation rate constant
            'M_sei': 0.162,     # SEI molar mass (kg/mol)
            'rho_sei': 1690,    # SEI density (kg/m^3)
            'kappa_sei': 1e-6,  # SEI ionic conductivity (S/m)
            'c_EC0': 4541,      # EC concentration in electrolyte (mol/m^3)
            
            # Thermal
            'h': 5.0,           # heat transfer coefficient (W/m^2/K)
            'rho_cp': 2.5e6,    # volumetric heat capacity (J/m^3/K)
            'T_amb': 298.15,    # ambient temperature (K)
            'R_cell': 0.02,     # cell ohmic resistance (Ohm)
        }
        if params is not None:
            self.params.update(params)
        self.p = self.params
        
        # Initialize state variables
        self.sei_thickness = 1e-9  # initial SEI thickness (m)
        self.T_cell = self.p['T_amb']
        
        # Compute derived quantities
        self._update_derived()
    
    def _update_derived(self):
        """Compute derived geometric quantities."""
        p = self.p
        p['a_p'] = 3 * p['eps_s_p'] / p['R_p']  # specific interfacial area cathode
        p['a_n'] = 3 * p['eps_s_n'] / p['R_n']  # specific interfacial area anode
    
    def sei_resistance(self):
        """SEI film resistance (Ohm-m^2)."""
        return self.sei_thickness / self.p['kappa_sei']
    
    def simulate_discharge(self, I_app, T_amb, t_end=None, n_points=200):
        """
        Simulate a constant-current discharge.
        
        Args:
            I_app: applied current (A), positive for discharge
            T_amb: ambient temperature (K)
            t_end: end time (s), auto-computed if None
            n_points: number of time points
        
        Returns:
            dict with t, V, T, Q, SOC arrays
        """
        p = self.p
        I_app = abs(I_app)  # discharge current magnitude
        I_density = I_app / p['A']  # current density (A/m^2)
        
        # Estimate theoretical capacity (Coulombs)
        Q_max_C = p['eps_s_p'] * p['L_p'] * p['A'] * p['c_s_max_p'] * F * (1.0 - p['x0'])
        
        if t_end is None:
            t_end = Q_max_C / I_app * 0.95  # 95% depth of discharge
        
        t = np.linspace(0, t_end, n_points)
        dt = t[1] - t[0]
        
        # Initialize states
        x = p['x0']  # cathode stoichiometry
        y = p['y0']  # anode stoichiometry
        T = T_amb
        sei = self.sei_thickness
        
        # Pre-allocate output arrays
        V_arr = np.zeros(n_points)
        T_arr = np.zeros(n_points)
        Q_arr = np.zeros(n_points)
        SOC_arr = np.zeros(n_points)
        sei_arr = np.zeros(n_points)
        
        # Surface area factors
        a_p_L_p = p['a_p'] * p['L_p']  # cathode surface/geometric area ratio
        a_n_L_n = p['a_n'] * p['L_n']  # anode surface/geometric area ratio
        
        # Capacity factors for stoichiometry update
        cap_factor_p = p['eps_s_p'] * p['L_p'] * p['c_s_max_p'] * F  # C/m^2
        cap_factor_n = p['eps_s_n'] * p['L_n'] * p['c_s_max_n'] * F  # C/m^2
        
        Q_discharged = 0.0  # Ah
        
        for i in range(n_points):
            # Surface concentrations (SP approximation: uniform)
            cs_p_surf = p['c_s_max_p'] * x
            cs_n_surf = p['c_s_max_n'] * y
            
            # Exchange current densities (A/m^2)
            # i0 = F * k * ce^0.5 * cs^0.5 * (cs_max - cs)^0.5
            sqrt_ce = np.sqrt(p['c_e'])
            i0_p = F * p['k_p'] * sqrt_ce * np.sqrt(max(cs_p_surf, 1e-6)) * np.sqrt(max(p['c_s_max_p'] - cs_p_surf, 1e-6))
            i0_n = F * p['k_n'] * sqrt_ce * np.sqrt(max(cs_n_surf, 1e-6)) * np.sqrt(max(p['c_s_max_n'] - cs_n_surf, 1e-6))
            
            # Butler-Volmer overpotentials
            # I_density = i0 * a * L * (exp(0.5*F*eta/RT) - exp(-0.5*F*eta/RT))
            # = 2 * i0 * a * L * sinh(0.5*F*eta/RT)
            # eta = RT/(0.5F) * arcsinh(I / (2*i0*a*L))
            arg_p = I_density / max(2 * a_p_L_p * i0_p, 1e-20)
            arg_n = I_density / max(2 * a_n_L_n * i0_n, 1e-20)
            
            # Clip to prevent numerical overflow
            arg_p = np.clip(arg_p, -1e6, 1e6)
            arg_n = np.clip(arg_n, -1e6, 1e6)
            
            eta_factor = R_gas * T / (0.5 * F)
            eta_p = eta_factor * np.arcsinh(arg_p)
            eta_n = eta_factor * np.arcsinh(arg_n)
            
            # Diffusion overpotentials (steady-state approximation)
            # eta_diff = RT/F * (I * R_p / (a * L * F * Ds * cs_max))
            eta_diff_p = R_gas * T / F * (I_density * p['R_p'] / max(a_p_L_p * F * p['D_s_p'] * p['c_s_max_p'], 1e-20))
            eta_diff_n = R_gas * T / F * (I_density * p['R_n'] / max(a_n_L_n * F * p['D_s_n'] * p['c_s_max_n'], 1e-20))
            
            # OCP
            U_p = ocp_nmc(x)
            U_n = ocp_graphite(y)
            
            # SEI resistance contribution
            R_sei = sei / p['kappa_sei']  # Ohm-m^2
            R_sei_total = R_sei / p['A']  # Ohm (divide by area to get resistance)
            
            # Cell voltage
            V_ocv = U_p - U_n
            V_cell = V_ocv - abs(eta_p) - abs(eta_n) - eta_diff_p - eta_diff_n - I_app * (p['R_cell'] + R_sei_total)
            V_cell = np.clip(V_cell, 2.0, 5.0)  # physical bounds
            
            # Heat generation (J/s)
            Q_gen = I_app * (V_ocv - V_cell)
            # Heat dissipation
            Q_diss = p['h'] * p['A'] * (T - T_amb)
            # Temperature update
            dT = (Q_gen - Q_diss) * dt / p['rho_cp']
            T += dT
            T = np.clip(T, T_amb - 5, T_amb + 50)
            
            # SEI growth (simplified Tafel kinetics)
            # The side reaction driving force is the anode overpotential
            phi_anode_ref = 0.0
            eta_sei = phi_anode_ref - U_n - I_app * R_sei_total
            # Side reaction current density (negative = reduction = SEI formation)
            i_sei = -F * p['k_sei'] * p['c_EC0'] * np.exp(-0.5 * F / (R_gas * T) * eta_sei)
            i_sei = np.clip(i_sei, -1e-3, 0.0)  # small loss
            
            # SEI thickness growth rate
            dsei_dt = -i_sei * p['M_sei'] / (2 * F * p['rho_sei'])
            sei += dsei_dt * dt
            sei = np.clip(sei, 1e-9, 1e-6)  # 1nm to 1um range
            
            # Update stoichiometry
            # During discharge: Li+ enters cathode (x INCREASES), leaves anode (y DECREASES)
            dx = I_density * dt / cap_factor_p
            dy = -I_density * dt / cap_factor_n
            
            x = np.clip(x + dx, 0.02, 0.98)
            y = np.clip(y + dy, 0.01, 0.99)
            
            Q_discharged += I_app * dt / 3600.0  # Ah
            
            # Store
            V_arr[i] = V_cell
            T_arr[i] = T
            Q_arr[i] = Q_discharged
            SOC_arr[i] = 1.0 - Q_discharged / (Q_max_C / 3600.0)
            sei_arr[i] = sei
        
        # Update instance state
        self.sei_thickness = sei
        self.T_cell = T
        
        return {
            't': t, 'V': V_arr, 'T': T_arr, 'Q': Q_arr, 'SOC': SOC_arr,
            'sei_thickness': sei
        }


# =============================================================================
# Latin Hypercube Sampling
# =============================================================================

def create_parameter_search_space():
    """Define the multi-dimensional parameter search space with bounds."""
    space = {
        'R_p':      (1e-6, 11e-6),        # cathode particle radius
        'R_n':      (1e-6, 11e-6),        # anode particle radius
        'k_p':      (1e-12, 1e-10),       # cathode reaction rate
        'k_n':      (1e-12, 1e-10),       # anode reaction rate
        'D_s_p':    (1e-14, 1e-12),       # cathode diffusivity
        'D_s_n':    (1e-15, 1e-13),       # anode diffusivity
        'eps_s_p':  (0.35, 0.55),         # cathode volume fraction
        'eps_s_n':  (0.40, 0.60),         # anode volume fraction
        'L_p':      (35e-6, 80e-6),       # cathode thickness
        'L_n':      (35e-6, 80e-6),       # anode thickness
        'k_sei':    (1e-14, 1e-12),       # SEI rate constant
        'h':        (1.0, 10.0),          # heat transfer coefficient
    }
    return space

def generate_lhs_samples(param_space, n_samples):
    """Generate Latin Hypercube samples in the parameter space."""
    param_names = list(param_space.keys())
    n_dims = len(param_names)
    
    sampler = qmc.LatinHypercube(d=n_dims, seed=42)
    samples_unit = sampler.random(n=n_samples)
    
    samples = {}
    for i, name in enumerate(param_names):
        lo, hi = param_space[name]
        if lo > 0 and hi / lo > 100:
            # Log-uniform sampling for wide ranges
            samples[name] = np.exp(
                np.log(lo) + samples_unit[:, i] * (np.log(hi) - np.log(lo))
            )
        else:
            samples[name] = lo + samples_unit[:, i] * (hi - lo)
    
    return samples, param_names


# =============================================================================
# Experiment data loaders
# =============================================================================

def load_nasa_data(battery_id=5):
    """Load NASA PCoE battery discharge data."""
    import scipy.io as sio
    filename = f'data/NASA PCoE Dataset Repository/1. BatteryAgingARC-FY08Q4/B{battery_id:04d}.mat'
    mat = sio.loadmat(filename)
    
    all_cycles = mat[f'B{battery_id:04d}']['cycle'][0, 0]
    discharge_data = []
    
    for i in range(len(all_cycles['type'])):
        cycle_type = str(all_cycles['type'][i])
        if 'discharge' in cycle_type:
            data = all_cycles['data'][0, i]
            d = {
                'cycle': i,
                'V': data['Voltage_measured'][0, 0].flatten(),
                'I': data['Current_measured'][0, 0].flatten(),
                'T': data['Temperature_measured'][0, 0].flatten(),
                't': data['Time'][0, 0].flatten(),
            }
            if 'Capacity' in data.dtype.names:
                cap = data['Capacity'][0, 0]
                if hasattr(cap, 'flatten'):
                    cap = cap.flatten()
                d['Capacity'] = float(cap[0]) if len(cap) > 0 else None
            discharge_data.append(d)
    
    return discharge_data

def load_cs2_data(filename='data/CS2_36/CS2_36_1_10_11.xlsx'):
    """Load CS2_36 cycle data from xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[1]]  # Channel sheet
    
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[5] is not None and row[6] is not None:
            rows.append({
                'cycle': row[5],
                'step': row[4],
                't': row[1],
                'I': row[6],
                'V': row[7],
                'Q_discharge': row[9] if row[9] is not None else 0,
            })
    
    # Group by cycle
    cycles = {}
    for r in rows:
        c = r['cycle']
        if c not in cycles:
            cycles[c] = {'t': [], 'V': [], 'I': [], 'Q': [], 'steps': set()}
        cycles[c]['t'].append(r['t'])
        cycles[c]['V'].append(r['V'])
        cycles[c]['I'].append(r['I'])
        cycles[c]['Q'].append(r['Q_discharge'])
        cycles[c]['steps'].add(r['step'])
    
    # Convert to arrays
    result = []
    for c in sorted(cycles.keys()):
        result.append({
            'cycle': c,
            't': np.array(cycles[c]['t']),
            'V': np.array(cycles[c]['V']),
            'I': np.array(cycles[c]['I']),
            'Q': np.array(cycles[c]['Q']),
            'steps': cycles[c]['steps']
        })
    
    return result

def load_oxford_data():
    """Load Oxford Battery Degradation Dataset."""
    import scipy.io as sio
    mat = sio.loadmat('data/Oxford Battery Degradation Dataset/ExampleDC_C1.mat')
    
    dc = mat['ExampleDC_C1'][0, 0]['dc'][0, 0]
    ch = mat['ExampleDC_C1'][0, 0]['ch'][0, 0]
    
    result = {
        'discharge': {
            't': dc['t'].flatten(),
            'V': dc['v'].flatten(),
            'I': dc['i'].flatten(),
            'Q': dc['q'].flatten(),
            'T': dc['T'].flatten(),
        },
        'charge': {
            't': ch['t'].flatten(),
            'V': ch['v'].flatten(),
            'I': ch['i'].flatten(),
            'Q': ch['q'].flatten(),
            'T': ch['T'].flatten(),
        }
    }
    return result


# =============================================================================
# ANN Meta-Model
# =============================================================================

class ANNMetaModel:
    """Simple feedforward neural network as meta-model surrogate."""
    
    def __init__(self, input_dim, hidden_dims=[64, 128, 64], output_dim=200):
        self.input_dim = input_dim
        self.hidden_dims = hidden_dims
        self.output_dim = output_dim
        self.weights = []
        self.biases = []
        self._build()
    
    def _build(self):
        dims = [self.input_dim] + self.hidden_dims + [self.output_dim]
        for i in range(len(dims) - 1):
            scale = np.sqrt(2.0 / (dims[i] + dims[i+1]))
            self.weights.append(np.random.randn(dims[i], dims[i+1]) * scale)
            self.biases.append(np.zeros(dims[i+1]))
    
    def forward(self, X):
        A = X
        for i in range(len(self.weights) - 1):
            Z = A @ self.weights[i] + self.biases[i]
            A = np.maximum(0, Z)
        Z = A @ self.weights[-1] + self.biases[-1]
        return Z
    
    def train(self, X, Y, epochs=500, lr=1e-3, batch_size=32, verbose=False):
        n_samples = X.shape[0]
        losses = []
        
        for epoch in range(epochs):
            idx = np.random.permutation(n_samples)
            X_shuffled = X[idx]
            Y_shuffled = Y[idx]
            
            total_loss = 0
            n_batches = 0
            
            for start in range(0, n_samples, batch_size):
                end = min(start + batch_size, n_samples)
                X_batch = X_shuffled[start:end]
                Y_batch = Y_shuffled[start:end]
                m = X_batch.shape[0]
                
                # Forward pass
                activations = [X_batch]
                Zs = []
                A = X_batch
                for i in range(len(self.weights) - 1):
                    Z = A @ self.weights[i] + self.biases[i]
                    Zs.append(Z)
                    A = np.maximum(0, Z)
                    activations.append(A)
                Z = A @ self.weights[-1] + self.biases[-1]
                Zs.append(Z)
                output = Z
                
                # MSE loss
                error = output - Y_batch
                loss = np.mean(error ** 2)
                
                # Skip batch if loss is NaN
                if np.isnan(loss) or np.isinf(loss):
                    continue
                    
                total_loss += loss
                n_batches += 1
                
                # Backprop
                dZ = 2 * error / m
                for i in reversed(range(len(self.weights))):
                    dW = activations[i].T @ dZ
                    db = np.sum(dZ, axis=0)
                    if i > 0:
                        dA = dZ @ self.weights[i].T
                        dZ = dA * (Zs[i-1] > 0).astype(float)
                    
                    # Gradient clipping
                    dW = np.clip(dW, -1.0, 1.0)
                    db = np.clip(db, -1.0, 1.0)
                    
                    self.weights[i] -= lr * dW
                    self.biases[i] -= lr * db
            
            if n_batches > 0:
                avg_loss = total_loss / n_batches
            else:
                avg_loss = float('nan')
            losses.append(avg_loss)
            
            if verbose and epoch % 100 == 0:
                print(f"  Epoch {epoch}: loss = {avg_loss:.6f}")
        
        return losses
    
    def predict(self, X):
        return self.forward(X)


# =============================================================================
# Genetic Algorithm
# =============================================================================

class GeneticAlgorithm:
    """Genetic Algorithm for parameter identification using ANN meta-model."""
    
    def __init__(self, param_space, ann_model, target_curve, population_size=100,
                 n_generations=50, mutation_rate=0.1, crossover_rate=0.8):
        self.param_space = param_space
        self.param_names = list(param_space.keys())
        self.n_params = len(self.param_names)
        self.ann = ann_model
        self.target = target_curve
        self.pop_size = population_size
        self.n_gen = n_generations
        self.mutation_rate = mutation_rate
        self.crossover_rate = crossover_rate
        
        self.lo = np.array([param_space[n][0] for n in self.param_names])
        self.hi = np.array([param_space[n][1] for n in self.param_names])
        self.is_log = np.array([(self.lo[i] > 0 and self.hi[i]/self.lo[i] > 100) 
                                for i in range(self.n_params)])
    
    def _decode(self, encoded):
        params = np.zeros(self.n_params)
        for i in range(self.n_params):
            if self.is_log[i]:
                params[i] = np.exp(np.log(self.lo[i]) + encoded[i] * (np.log(self.hi[i]) - np.log(self.lo[i])))
            else:
                params[i] = self.lo[i] + encoded[i] * (self.hi[i] - self.lo[i])
        return params
    
    def _fitness(self, encoded_individual):
        params = self._decode(encoded_individual)
        norm_params = (params - self.lo) / (self.hi - self.lo + 1e-12)
        predicted = self.ann.predict(norm_params.reshape(1, -1)).flatten()
        
        # Interpolate to match target length
        target_len = len(self.target)
        if len(predicted) != target_len:
            x_old = np.linspace(0, 1, len(predicted))
            x_new = np.linspace(0, 1, target_len)
            predicted = np.interp(x_new, x_old, predicted)
        
        mse = np.mean((predicted - self.target) ** 2)
        if np.isnan(mse) or np.isinf(mse):
            return 1e10
        return float(mse)
    
    def _initialize_population(self):
        sampler = qmc.LatinHypercube(d=self.n_params, seed=123)
        return sampler.random(n=self.pop_size)
    
    def _tournament_select(self, population, fitnesses, k=3):
        n = len(population)
        selected = []
        valid_indices = np.where(np.isfinite(fitnesses))[0]
        if len(valid_indices) == 0:
            # All bad - random select
            for _ in range(n):
                selected.append(population[np.random.randint(n)].copy())
            return np.array(selected)
        
        for _ in range(n):
            candidates = np.random.choice(valid_indices, min(k, len(valid_indices)), replace=False)
            winner = candidates[np.argmin([fitnesses[c] for c in candidates])]
            selected.append(population[winner].copy())
        return np.array(selected)
    
    def _crossover(self, parent1, parent2):
        if np.random.random() > self.crossover_rate:
            return parent1.copy(), parent2.copy()
        
        u = np.random.random(self.n_params)
        beta = np.where(u <= 0.5, (2*u)**(1/3), (1/(2*(1-u)))**(1/3))
        child1 = 0.5 * ((1+beta)*parent1 + (1-beta)*parent2)
        child2 = 0.5 * ((1-beta)*parent1 + (1+beta)*parent2)
        return np.clip(child1, 0, 1), np.clip(child2, 0, 1)
    
    def _mutate(self, individual):
        mutant = individual.copy()
        for i in range(self.n_params):
            if np.random.random() < self.mutation_rate:
                u = np.random.random()
                delta = (2*u)**(1/4) - 1 if u < 0.5 else 1 - (2*(1-u))**(1/4)
                mutant[i] += delta
        return np.clip(mutant, 0, 1)
    
    def run(self, verbose=True):
        pop = self._initialize_population()
        best_fitness_history = []
        best_individual = None
        best_fitness = float('inf')
        
        for gen in range(self.n_gen):
            fitnesses = np.array([self._fitness(ind) for ind in pop])
            
            gen_best_idx = np.argmin(fitnesses)
            if fitnesses[gen_best_idx] < best_fitness:
                best_fitness = fitnesses[gen_best_idx]
                best_individual = pop[gen_best_idx].copy()
            
            best_fitness_history.append(best_fitness)
            
            if verbose and gen % 10 == 0:
                print(f"  Gen {gen:3d}: best fitness = {best_fitness:.6f}")
            
            selected = self._tournament_select(pop, fitnesses)
            
            next_pop = []
            for i in range(0, self.pop_size, 2):
                p1, p2 = selected[i], selected[min(i+1, self.pop_size-1)]
                c1, c2 = self._crossover(p1, p2)
                c1 = self._mutate(c1)
                c2 = self._mutate(c2)
                next_pop.append(c1)
                next_pop.append(c2)
            
            if best_individual is not None:
                next_pop[0] = best_individual.copy()
            pop = np.array(next_pop[:self.pop_size])
        
        if best_individual is None:
            # Return center of search space as fallback
            best_individual = np.ones(self.n_params) * 0.5
            best_fitness = float('inf')
        
        best_params = self._decode(best_individual)
        best_params_dict = {self.param_names[i]: float(best_params[i]) for i in range(self.n_params)}
        
        return best_params_dict, best_fitness, best_fitness_history


# =============================================================================
# Module test
# =============================================================================

if __name__ == '__main__':
    print("=" * 70)
    print("MMGA: Meta-Model based Genetic Algorithm for ECAT Parameter Identification")
    print("=" * 70)
    
    # Test the model
    print("\nTesting ECAT model...")
    model = ECATModel()
    result = model.simulate_discharge(2.0, 298.15, t_end=3600, n_points=10)
    print(f"  Voltage range: [{result['V'].min():.3f}, {result['V'].max():.3f}] V")
    print(f"  Temperature: [{result['T'].min():.1f}, {result['T'].max():.1f}] K")
    print(f"  Capacity: {result['Q'][-1]:.3f} Ah")
    print(f"  SEI thickness: {result['sei_thickness']:.4e} m")
    
    print("\nModule loaded successfully.")
